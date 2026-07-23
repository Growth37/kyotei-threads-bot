#!/usr/bin/env python3
"""投稿の表示回数トラッキングと的中チェック.

- posts_log.json に記録された各投稿について
  1時間後 / 5時間後 / 1日後 の表示回数(views)をThreads APIから取得
- レース結果(Boatrace Open API)と照合し、6点の中に3連単の結果があれば
  的中投稿を自動でThreadsに出す
- 結果を metrics.csv / metrics.xlsx にまとめる

環境変数:
  THREADS_ACCESS_TOKEN Threads長期アクセストークン
"""

import csv
import json
import os
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

JST = timezone(timedelta(hours=9))
THREADS_API = "https://graph.threads.net/v1.0"
RESULTS_URL = "https://boatraceopenapi.github.io/results/v2/{y}/{ymd}.json"

LOG_FILE = "posts_log.json"
CSV_FILE = "metrics.csv"
XLSX_FILE = "metrics.xlsx"

SNAPSHOTS = [("1h", 1.0), ("5h", 5.0), ("24h", 24.0)]


def http_get_json(url: str):
    req = urllib.request.Request(url, headers={"User-Agent": "kyotei-bot/1.0"})
    with urllib.request.urlopen(req, timeout=30) as res:
        return json.loads(res.read().decode("utf-8"))


def http_post(url: str, params: dict):
    data = urllib.parse.urlencode(params).encode("utf-8")
    req = urllib.request.Request(url, data=data, method="POST")
    with urllib.request.urlopen(req, timeout=30) as res:
        return json.loads(res.read().decode("utf-8"))


def fetch_views(post_id: str, token: str):
    """投稿の表示回数(views)を取得。失敗時はNone."""
    try:
        qs = urllib.parse.urlencode({"metric": "views", "access_token": token})
        data = http_get_json(f"{THREADS_API}/{post_id}/insights?{qs}")
        for m in data.get("data") or []:
            if m.get("name") != "views":
                continue
            if isinstance(m.get("total_value"), dict) and "value" in m["total_value"]:
                return int(m["total_value"]["value"])
            values = m.get("values") or []
            if values and "value" in values[0]:
                return int(values[0]["value"])
    except Exception as e:  # noqa: BLE001
        print(f"  views取得失敗 ({post_id}): {e}")
    return None


_results_cache = {}


def fetch_results(race_date: str):
    """指定日のレース結果一覧を取得 (日付ごとにキャッシュ)."""
    if race_date in _results_cache:
        return _results_cache[race_date]
    ymd = race_date.replace("-", "")
    try:
        data = http_get_json(RESULTS_URL.format(y=ymd[:4], ymd=ymd))
        _results_cache[race_date] = data.get("results") or []
    except Exception as e:  # noqa: BLE001
        print(f"  結果取得失敗 ({race_date}): {e}")
        _results_cache[race_date] = []
    return _results_cache[race_date]


def find_race_result(entry: dict):
    """該当レースの結果から (3連単の組番, 配当) を返す。未確定ならNone."""
    for race in fetch_results(entry["race_date"]):
        if int(race.get("race_stadium_number") or 0) != int(entry["stadium_number"]):
            continue
        if int(race.get("race_number") or 0) != int(entry["race_number"]):
            continue
        places = {}
        for b in race.get("boats") or []:
            p = b.get("racer_place_number")
            if p in (1, 2, 3):
                places[int(p)] = int(b["racer_boat_number"])
        if len(places) < 3:
            return None  # まだ確定していない/欠場等
        combo = f"{places[1]}-{places[2]}-{places[3]}"
        payout = None
        tri = (race.get("payouts") or {}).get("trifecta") or []
        for t in tri:
            if t.get("combination") == combo:
                payout = t.get("payout")
                break
        return combo, payout
    return None


def get_user_id(token: str) -> str:
    qs = urllib.parse.urlencode({"fields": "id", "access_token": token})
    me = http_get_json(f"{THREADS_API}/me?{qs}")
    return me["id"]


def current_streak(log: list) -> int:
    """結果確定済みレースを時系列に並べ、直近から遡った連続的中数を返す."""
    resolved = [e for e in log if e.get("result")]
    resolved.sort(key=lambda x: x.get("race_closed_at") or "")
    streak = 0
    for e in reversed(resolved):
        if e.get("hit"):
            streak += 1
        else:
            break
    return streak


def fetch_permalink(post_id: str, token: str):
    """投稿のURL(permalink)を取得。失敗時はNone."""
    try:
        qs = urllib.parse.urlencode({"fields": "permalink", "access_token": token})
        data = http_get_json(f"{THREADS_API}/{post_id}?{qs}")
        return data.get("permalink")
    except Exception as e:  # noqa: BLE001
        print(f"  permalink取得失敗 ({post_id}): {e}")
        return None


def post_hit(entry: dict, payout, token: str, user_id: str, streak: int = 1) -> bool:
    lines = []
    if streak >= 2:
        lines.append(f"🔥{streak}連勝中やで!!")
    lines += [
        "🎯的中や!!",
        f"{entry['stadium']}{entry['race_number']}R 3連単 {entry['result']}",
    ]
    if payout:
        lines.append(f"配当 {int(payout):,}円")
    if entry.get("nerai") and entry["result"] in entry["nerai"]:
        lines.append("狙いの中穴がズバッとハマったわ🎯")
    permalink = fetch_permalink(entry["post_id"], token)
    if permalink:
        lines += ["", "👇この予想やで", permalink]
    lines += ["", "※舟券は自己責任でな🙏", "#競艇 #ボートレース #競艇予想"]
    text = "\n".join(lines)
    try:
        c = http_post(
            f"{THREADS_API}/{user_id}/threads",
            {"media_type": "TEXT", "text": text, "access_token": token},
        )
        if not c.get("id"):
            print(f"  的中投稿コンテナ失敗: {c}")
            return False
        time.sleep(35)
        r = http_post(
            f"{THREADS_API}/{user_id}/threads_publish",
            {"creation_id": c["id"], "access_token": token},
        )
        if r.get("id"):
            print(f"  的中投稿完了! post id = {r['id']}")
            return True
        print(f"  的中投稿公開失敗: {r}")
    except Exception as e:  # noqa: BLE001
        print(f"  的中投稿エラー: {e}")
    return False


def write_csv(log: list):
    headers = [
        "投稿日", "投稿時刻", "場", "R", "締切", "予想6点",
        "結果(3連単)", "的中", "配当",
        "views 1時間後", "views 5時間後", "views 1日後",
    ]
    rows = []
    for e in sorted(log, key=lambda x: x["posted_at"]):
        views = e.get("views") or {}
        hit = ""
        if e.get("result"):
            hit = "◎的中" if e.get("hit") else "×"
        rows.append([
            e["posted_at"][:10], e["posted_at"][11:16],
            e["stadium"], e["race_number"], e["race_closed_at"][11:16],
            " / ".join(e.get("combos") or []),
            e.get("result") or "", hit,
            e.get("payout") or "",
            views.get("1h", ""), views.get("5h", ""), views.get("24h", ""),
        ])
    with open(CSV_FILE, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "投稿実績"
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
        for r in rows:
            ws.append(r)
        widths = [11, 9, 8, 5, 8, 40, 12, 8, 10, 13, 13, 13]
        for i, wd in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wd
        wb.save(XLSX_FILE)
    except Exception as e:  # noqa: BLE001
        print(f"xlsx生成スキップ: {e}")


def main():
    token = os.environ.get("THREADS_ACCESS_TOKEN", "").strip()
    if not token:
        print("THREADS_ACCESS_TOKEN が未設定です。", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(LOG_FILE):
        print("posts_log.json がまだありません。投稿後に生成されます。")
        return
    with open(LOG_FILE, encoding="utf-8") as f:
        log = json.load(f)
    if not log:
        print("記録された投稿がまだありません。")
        write_csv(log)
        return

    now = datetime.now(JST)
    user_id = None
    changed = False

    for e in log:
        posted = datetime.fromisoformat(e["posted_at"])
        elapsed_h = (now - posted).total_seconds() / 3600
        e.setdefault("views", {})

        # 表示回数スナップショット
        for key, hours in SNAPSHOTS:
            if key not in e["views"] and elapsed_h >= hours:
                v = fetch_views(e["post_id"], token)
                if v is not None:
                    e["views"][key] = v
                    e.setdefault("views_at", {})[key] = round(elapsed_h, 2)
                    print(f"{e['stadium']}{e['race_number']}R: {key} views = {v}")
                    changed = True

        # 的中チェック (締切40分後以降)
        if not e.get("result"):
            closed = datetime.strptime(
                e["race_closed_at"], "%Y-%m-%d %H:%M:%S"
            ).replace(tzinfo=JST)
            if now >= closed + timedelta(minutes=40):
                res = find_race_result(e)
                if res:
                    combo, payout = res
                    e["result"] = combo
                    e["payout"] = payout
                    e["hit"] = combo in (e.get("combos") or [])
                    changed = True
                    mark = "🎯的中!" if e["hit"] else "不的中"
                    print(f"{e['stadium']}{e['race_number']}R 結果 {combo} → {mark}")
                    if e["hit"] and not e.get("hit_posted"):
                        if user_id is None:
                            user_id = get_user_id(token)
                        if post_hit(e, payout, token, user_id, current_streak(log)):
                            e["hit_posted"] = True

    if changed:
        with open(LOG_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=1)
    write_csv(log)
    print("完了。")


if __name__ == "__main__":
    main()
